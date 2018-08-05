# _*_ coding=utf-8 _*_

from openpyxl import load_workbook

def stu_num(file_name):
    # 定表名
    wb = load_workbook(file_name) # 载入Excel文件
    sheet_name = wb.sheetnames # 获取所有的表名称（数组）
    sheet_name_first = str(sheet_name[0]) # 获取第一个表名称
    sheet = wb[sheet_name_first] # 载入第一个表的内容
    target1 = [] # 定位学号数组
    target2 = [] # 定位姓名数组
    stu_num_nam_error = [] # 学号姓名报错三维数组
    errors = [] # 报错信息数组
    # 定位
    for i in range(1, 5):
        for j in range(1, 5):
            cell_value = sheet.cell(row=i, column=j).value # 寻找"学号"进而定位
            if cell_value == 'XH':
                target1.append(i)
                target1.append(j) # 把"学号"的位置传入数组target1
                break # 跳出里层循环
        else:
            continue
        break # 跳出最外层循环

    #'''	
    for i in range(1, 5):
        for j in range(1, 5):
            cell_value = sheet.cell(row=i, column=j).value # 寻找"姓名"进而定位
            if cell_value == 'XM':
                target2.append(i)
                target2.append(j) # 把"姓名"的位置传入数组target2
                break # 跳出里层循环
        else:
            continue
        break # 跳出最外层循环
    #'''

    # 导入数组
    first_student_number = target1[0] + 1 # 第一行学号
    student_number_colms = target1[1] # 学号的列
    first_student_name = target2[0] + 1 # 第一行学号
    student_name_colms = target2[1] # 学号列
    maxraw = sheet.max_row + 1 # 获取最大行数
    student_number = [] # 学号的数组
	#student_name = [] # 姓名的数组
    student_name = []
    check = 0 # 记录空元素的数量
    if first_student_name != first_student_number: # 判断学号列和姓名列是否对齐
        errors.append('file error,please check')
    else:
        for i in range(first_student_number,maxraw):
            if str(sheet.cell(row = i,column = student_number_colms).value) != 'None' and str(sheet.cell(row = i,column = student_name_colms).value) != 'None':
                # 判断学号姓名是否为空，均不为空则分别添加到两个数组里
                student_number.append(str(sheet.cell(row = i,column = student_number_colms).value)) # 学号数组
                student_name.append(str(sheet.cell(row = i,column = student_name_colms).value)) # 姓名数组
                check = 0 # 清空检测为空变量
            elif str(sheet.cell(row = i,column = student_number_colms).value) == 'None' and str(sheet.cell(row = i,column = student_name_colms).value) == 'None':
                # 判断学号姓名是否为空，为空则不写入。若连续三个均为空则跳出循环
                check += 1
                if check <3 :
                    continue
                else:
                    break
            else:
                # 若学号姓名有一为空，则写入error数组
                errors.append(str(i)+'error')
    
    #三位数组写入
    len_stu_name = len(student_name) # 求姓名数组长度
    len_stu_number = int(len(student_number)) # 求学号数组长度
    len_error = len(errors) # 求err咋or长度
    stu_num_nam = [] # 学号姓名对应数组
    stu_num_nam_error.append(errors) # 添加error到学号姓名报错数组中

    for i in range(0,len_stu_number):
        stun = student_number[i] # 第i位学号存入stun
        stua = student_name[i] # 第i位姓名存入stua
        stu_num_nam.append(stun) # 第i位学号填入学号姓名对应数组
        stu_num_nam.append(stua) # 第i位姓名填入学号姓名对应数组
        stu_num_nam_error.append(stu_num_nam) # 学号姓名数组加入到学号姓名报错数组
        stu_num_nam = [] # 清空学号姓名数组
    
    return stu_num_nam_error
    '''
    逻辑如下：
        先定位学号和姓名位置，保证无论学号列在前还是姓名列在前都可以纵向读出
        然后判断是否对齐。如果表头未对齐，则报错并不写入
        再判断每一个学号是否对应一个姓名，是否每一个姓名都对应一个学号。对应则填入，不对应就写入报错信息
        将报错数组填入总返回数组
        读取第i位的学号和姓名（上几步保证一一对应）写入数组，并将此数组填入总返回数组
        返回数组
    '''

def active_codes(file_name):
    # 定表名
    wb = load_workbook(file_name) # 载入Excel文件
    sheet_name = wb.sheetnames # 获取所有的表名称（数组）
    sheet_name_first = str(sheet_name[0]) # 获取第一个表名称
    sheet = wb[sheet_name_first] # 载入第一个表的内容
    target = [] # 全局数组

    # 定位
    for i in range(1, 5):
        for j in range(1, 5):
            cell_value = sheet.cell(row=i, column=j).value # 寻找"激活码"进而定位
            if cell_value == '激活码':
                target.append(i)
                target.append(j) # 把"激活码"的位置传入数组e
                break # 跳出里层循环
        else:
            continue
        break # 跳出最外层循环
    
    # 导入数组
    first_code = target[0]+1 # 第一行激活码
    colm = target[1] # 激活码的列
    maxraw = sheet.max_row + 1# 取得最大行数
    active_code = [] # 激活码的数组
    check = 0 # 记录空元素的数量
    for i in range(first_code, maxraw):
        if str(sheet.cell(row=i, column=colm).value) != 'None': # 判断是否是空，空值则+1，否则计入激活码数组
            active_code.append(str(sheet.cell(row=i, column=colm).value))
            check = 0 # 将记录空元素数量的变量归零
        else:
            check += 1
            if check <= 3: # 如果连续四个空值则跳出循环，否则继续循环
                continue
            else:
                break
    
    return active_code
    '''
    逻辑如下：
        取第一个表，定位"激活码"
        逐个纵向读取激活码并判断是否为空，非空则计入返回数组，为空则累计，连续四个空值跳出循环
    '''